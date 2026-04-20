#!/usr/bin/env python3
"""
Export seat-counts.csv to seat-counts.xlsx.

This keeps a human-friendly workbook in sync with the append-only CSV that the
pipeline writes during Phase 2.
"""

import csv
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
SEAT_CSV = DATA_DIR / "seat-counts.csv"
SEAT_XLSX = DATA_DIR / "seat-counts.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _autowidth(ws, sample_rows=200):
    max_row = min(ws.max_row, sample_rows)
    for col_idx in range(1, ws.max_column + 1):
        values = []
        for row_idx in range(1, max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            values.append("" if value is None else str(value))
        width = min(max((len(v) for v in values), default=10) + 2, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _format_pct_column(ws, headers, column_name):
    if column_name not in headers:
        return
    col_idx = headers.index(column_name) + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value in ("", None):
            continue
        try:
            cell.value = float(cell.value) / 100.0
            cell.number_format = "0.0%"
        except (TypeError, ValueError):
            continue


def load_rows():
    if not SEAT_CSV.exists():
        return [], []
    with SEAT_CSV.open("r", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = reader.fieldnames or []
    return headers, rows


def build_workbook(headers, rows):
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "Seat Counts Export"
    summary["A1"].font = Font(bold=True, size=14)

    latest_weekend = max((row.get("weekend_of", "") for row in rows if row.get("weekend_of")), default="")
    latest_rows = [row for row in rows if row.get("weekend_of", "") == latest_weekend] if latest_weekend else rows
    actual_rows = [
        row for row in latest_rows
        if row.get("total_seats") and row.get("seats_sold") and row.get("occupancy_pct")
    ]

    summary["A3"] = "Latest weekend"
    summary["B3"] = latest_weekend or "N/A"
    summary["A4"] = "Total CSV rows"
    summary["B4"] = len(rows)
    summary["A5"] = "Latest-weekend rows"
    summary["B5"] = len(latest_rows)
    summary["A6"] = "Rows with seat counts"
    summary["B6"] = len(actual_rows)

    summary["A8"] = "Rows by timezone"
    summary["A9"] = "Timezone"
    summary["B9"] = "Rows"
    for cell in (summary["A9"], summary["B9"]):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, (tz, count) in enumerate(sorted(Counter(row.get("timezone", "") for row in latest_rows).items()), start=10):
        summary[f"A{idx}"] = tz or "?"
        summary[f"B{idx}"] = count
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 16

    raw = wb.create_sheet("Seat Counts Raw")
    if headers:
        raw.append(headers)
        for row in rows:
            raw.append([row.get(h, "") for h in headers])
        _style_header(raw)
        _format_pct_column(raw, headers, "occupancy_pct")
        _autowidth(raw)

    actuals = wb.create_sheet("Actual Seat Counts")
    if headers:
        actuals.append(headers)
        for row in actual_rows:
            actuals.append([row.get(h, "") for h in headers])
        _style_header(actuals)
        _format_pct_column(actuals, headers, "occupancy_pct")
        _autowidth(actuals)

    return wb


def main():
    headers, rows = load_rows()
    wb = build_workbook(headers, rows)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(SEAT_XLSX)
    print(f"Updated {SEAT_XLSX}")


if __name__ == "__main__":
    main()
