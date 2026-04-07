"""
Build 2026-04-05-box-office.xlsx
Filters seat-counts.csv to 2026-04-03 and 2026-04-04, requires total_seats.
4 Sheets: Summary by Movie, By Theatre, By Showtime, Raw
"""

import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

OUTPUT = "./box-office-tracker/data/2026-04-05-box-office.xlsx"
CSV = "./box-office-tracker/data/seat-counts.csv"
TARGET_DATES = {'2026-04-03', '2026-04-04'}

# ── Load & filter ─────────────────────────────────────────────────────────────
df_all = pd.read_csv(CSV, dtype=str, on_bad_lines='skip')
df_date = df_all[df_all['date'].isin(TARGET_DATES)].copy()
df = df_date[df_date['total_seats'].notna() & (df_date['total_seats'].str.strip() != '')].copy()

for col in ['total_seats', 'seats_sold', 'seats_available', 'occupancy_pct', 'minutes_after_showtime']:
    df[col] = pd.to_numeric(df[col], errors='coerce')

# ── Format priority helper ─────────────────────────────────────────────────────
def fmt_priority(fmt):
    if isinstance(fmt, str):
        if 'IMAX' in fmt:
            return 1
        if 'Dolby' in fmt:
            return 2
    return 3

df['_fmt_priority'] = df['auditorium_type'].apply(fmt_priority)

# ── Showtime sort helper ───────────────────────────────────────────────────────
def showtime_sort(st):
    if not isinstance(st, str):
        return 9999
    m = re.match(r'(\d+):(\d+)(am|pm)', st.strip().lower())
    if not m:
        return 9999
    h, mi, ap = int(m.group(1)), int(m.group(2)), m.group(3)
    if ap == 'pm' and h != 12:
        h += 12
    if ap == 'am' and h == 12:
        h = 0
    return h * 60 + mi

df['_showtime_sort'] = df['showtime'].apply(showtime_sort)

# ── TZ sort: ET before PT ──────────────────────────────────────────────────────
TZ_ORDER = {'ET': 0, 'CT': 1, 'MT': 2, 'PT': 3}
df['_tz_sort'] = df['timezone'].map(TZ_ORDER).fillna(9)

# ── Styling constants ──────────────────────────────────────────────────────────
DARK_BLUE = "1F3864"
WHITE = "FFFFFF"
ALT_ROW = "EBF2FA"

header_font = Font(name='Arial', bold=True, color=WHITE, size=11)
header_fill = PatternFill('solid', fgColor=DARK_BLUE)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

body_font = Font(name='Arial', size=10)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')
thin_border = Border(bottom=Side(style='thin', color='CCCCCC'))

green_fill   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
green_font   = Font(name='Arial', size=10, color="276221")
yellow_font  = Font(name='Arial', size=10, color="9C6500")
red_font     = Font(name='Arial', size=10, color="9C0006")

def style_header_row(ws, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

def style_body_cell(cell, alt=False, align='left', num_fmt=None,
                    green=False, yellow=False, red=False):
    font = Font(name='Arial', size=10,
                color=("276221" if green else "9C6500" if yellow else "9C0006" if red else "000000"))
    cell.font = font
    cell.alignment = center_align if align == 'center' else left_align
    cell.border = thin_border
    if green:
        cell.fill = green_fill
    elif yellow:
        cell.fill = yellow_fill
    elif red:
        cell.fill = red_fill
    elif alt:
        cell.fill = PatternFill('solid', fgColor=ALT_ROW)
    if num_fmt:
        cell.number_format = num_fmt

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def write_headers(ws, headers):
    ws.row_dimensions[1].height = 32
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i).value = h
    style_header_row(ws, len(headers))
    ws.freeze_panes = 'A2'

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Summary by Movie
# Columns: Movie, Date, Day of Week, TZ, # Theatres, # Showtime Records,
#          Avg Occupancy %, Max Occupancy %, Min Occupancy %
# Sort: Movie, Date, TZ
# ══════════════════════════════════════════════════════════════════════════════
summary_rows = []
for (movie, date, tz), grp in df.groupby(['movie_title', 'date', 'timezone']):
    day_of_week = grp['day_of_week'].iloc[0]
    n_theatres = grp['theatre_name'].nunique()
    n_records = len(grp)
    occ = grp['occupancy_pct'].dropna()
    avg_occ = round(float(occ.mean()), 1) if len(occ) else None
    max_occ = round(float(occ.max()), 1) if len(occ) else None
    min_occ = round(float(occ.min()), 1) if len(occ) else None
    summary_rows.append({
        'Movie': movie,
        'Date': date,
        'Day of Week': day_of_week,
        'TZ': tz,
        '# Theatres': n_theatres,
        '# Showtime Records': n_records,
        'Avg Occupancy %': avg_occ,
        'Max Occupancy %': max_occ,
        'Min Occupancy %': min_occ,
        '_tz_sort': TZ_ORDER.get(tz, 9),
    })
df_summary = (pd.DataFrame(summary_rows)
              .sort_values(['Movie', 'Date', '_tz_sort'])
              .reset_index(drop=True))

SUM_HEADERS = ['Movie', 'Date', 'Day of Week', 'TZ', '# Theatres',
               '# Showtime Records', 'Avg Occupancy %', 'Max Occupancy %', 'Min Occupancy %']
SUM_WIDTHS  = [36, 14, 14, 6, 14, 20, 18, 18, 18]

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — By Theatre
# One row per theatre+movie+date (best occupancy row per theatre per day)
# Columns: Movie, Date, TZ, Theatre Name, City, Top Format, Best Occupancy %,
#          # Formats Scraped, # Showtime Records, AMC Seat Map URL
# Sort: Date, TZ, Theatre Name
# ══════════════════════════════════════════════════════════════════════════════
theatre_rows = []
for (movie, date, theatre, city, tz), grp in df.groupby(
        ['movie_title', 'date', 'theatre_name', 'theatre_city', 'timezone']):
    occ = grp['occupancy_pct'].dropna()
    best_occ = round(float(occ.max()), 1) if len(occ) else None
    # Top format = format of row with best occupancy (or highest priority if tie)
    if len(occ):
        best_idx = grp['occupancy_pct'].idxmax()
        top_format = grp.loc[best_idx, 'auditorium_type']
    else:
        top_format = grp.sort_values('_fmt_priority').iloc[0]['auditorium_type']
    n_formats = grp['auditorium_type'].nunique()
    n_records = len(grp)
    theatre_rows.append({
        'Movie': movie,
        'Date': date,
        'TZ': tz,
        'Theatre Name': theatre,
        'City': city,
        'Top Format': top_format,
        'Best Occupancy %': best_occ,
        '# Formats Scraped': n_formats,
        '# Showtime Records': n_records,
        'AMC Seat Map URL': None,
        '_tz_sort': TZ_ORDER.get(tz, 9),
    })
df_theatre = (pd.DataFrame(theatre_rows)
              .sort_values(['Date', '_tz_sort', 'Theatre Name'])
              .reset_index(drop=True))

TH_HEADERS = ['Movie', 'Date', 'TZ', 'Theatre Name', 'City', 'Top Format',
              'Best Occupancy %', '# Formats Scraped', '# Showtime Records', 'AMC Seat Map URL']
TH_WIDTHS  = [36, 14, 6, 36, 18, 28, 18, 18, 20, 36]

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — By Showtime
# One row per theatre+movie+format+showtime
# Sort: Date, TZ, Theatre, Format priority, Showtime
# ══════════════════════════════════════════════════════════════════════════════
df_byshow = (df.sort_values(['date', '_tz_sort', 'theatre_name', '_fmt_priority', '_showtime_sort'])
             .reset_index(drop=True))

SHOW_CSV_COLS = ['movie_title', 'date', 'day_of_week', 'timezone', 'theatre_name',
                 'theatre_city', 'auditorium_type', 'showtime', 'check_time',
                 'minutes_after_showtime', 'total_seats', 'seats_sold',
                 'seats_available', 'occupancy_pct']
SHOW_HEADERS = ['Movie', 'Date', 'Day', 'TZ', 'Theatre', 'City', 'Format',
                'Showtime', 'Check Time', 'Minutes After Showtime',
                'Total Seats', 'Seats Sold', 'Seats Available', 'Occupancy %',
                'AMC Seat Map URL']
SHOW_WIDTHS  = [36, 12, 12, 6, 36, 18, 28, 10, 24, 22, 12, 12, 16, 14, 36]

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Raw
# ══════════════════════════════════════════════════════════════════════════════
raw_cols = [c for c in df_date.columns if not c.startswith('_')]
df_raw = df_date[raw_cols].copy()

# ══════════════════════════════════════════════════════════════════════════════
# Build workbook
# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ── Sheet 1: Summary by Movie ─────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Summary by Movie'
write_headers(ws1, SUM_HEADERS)
for i, w in enumerate(SUM_WIDTHS, 1):
    set_col_width(ws1, i, w)

for r_idx, row in df_summary.iterrows():
    excel_row = r_idx + 2
    alt = r_idx % 2 == 1
    for c_idx, col in enumerate(SUM_HEADERS, 1):
        val = row[col] if col in row else None
        if isinstance(val, float) and pd.isna(val):
            val = None
        cell = ws1.cell(row=excel_row, column=c_idx, value=val)
        if col in ('Avg Occupancy %', 'Max Occupancy %', 'Min Occupancy %'):
            style_body_cell(cell, alt=alt, align='center', num_fmt='0.0"%"')
        elif col in ('# Theatres', '# Showtime Records'):
            style_body_cell(cell, alt=alt, align='center')
        elif col in ('Date', 'Day of Week', 'TZ'):
            style_body_cell(cell, alt=alt, align='center')
        else:
            style_body_cell(cell, alt=alt, align='left')

# ── Sheet 2: By Theatre ───────────────────────────────────────────────────────
ws2 = wb.create_sheet('By Theatre')
write_headers(ws2, TH_HEADERS)
for i, w in enumerate(TH_WIDTHS, 1):
    set_col_width(ws2, i, w)

for r_idx, row in df_theatre.iterrows():
    excel_row = r_idx + 2
    alt = r_idx % 2 == 1
    for c_idx, col in enumerate(TH_HEADERS, 1):
        val = row.get(col, None)
        if isinstance(val, float) and pd.isna(val):
            val = None
        cell = ws2.cell(row=excel_row, column=c_idx, value=val)
        if col == 'AMC Seat Map URL':
            if val:
                cell.hyperlink = val
                cell.font = Font(name='Arial', size=10, color="0563C1", underline='single')
                cell.border = thin_border
            else:
                style_body_cell(cell, alt=alt, align='left')
        elif col == 'Best Occupancy %':
            style_body_cell(cell, alt=alt, align='center', num_fmt='0.0"%"')
        elif col in ('# Formats Scraped', '# Showtime Records'):
            style_body_cell(cell, alt=alt, align='center')
        elif col in ('Date', 'TZ'):
            style_body_cell(cell, alt=alt, align='center')
        else:
            style_body_cell(cell, alt=alt, align='left')

# ── Sheet 3: By Showtime ──────────────────────────────────────────────────────
ws3 = wb.create_sheet('By Showtime')
write_headers(ws3, SHOW_HEADERS)
for i, w in enumerate(SHOW_WIDTHS, 1):
    set_col_width(ws3, i, w)

occ_col_idx = SHOW_HEADERS.index('Occupancy %') + 1
occ_col_letter = get_column_letter(occ_col_idx)

for r_idx, row in df_byshow.iterrows():
    excel_row = r_idx + 2
    alt = r_idx % 2 == 1

    # Determine occupancy color
    occ_val = row['occupancy_pct']
    is_green = is_yellow = is_red = False
    if pd.notna(occ_val):
        if occ_val >= 70:
            is_green = True
        elif occ_val >= 40:
            is_yellow = True
        else:
            is_red = True

    for c_idx, (csv_col, hdr) in enumerate(zip(SHOW_CSV_COLS, SHOW_HEADERS[:len(SHOW_CSV_COLS)]), 1):
        val = row[csv_col]
        if isinstance(val, float) and pd.isna(val):
            val = None
        cell = ws3.cell(row=excel_row, column=c_idx, value=val)
        if hdr == 'Occupancy %':
            style_body_cell(cell, alt=alt, align='center', num_fmt='0.0"%"',
                            green=is_green, yellow=is_yellow, red=is_red)
        elif hdr in ('Total Seats', 'Seats Sold', 'Seats Available'):
            style_body_cell(cell, alt=alt, align='center', num_fmt='#,##0')
        elif hdr == 'Minutes After Showtime':
            style_body_cell(cell, alt=alt, align='center', num_fmt='#,##0')
        elif hdr in ('Date', 'Day', 'TZ', 'Showtime'):
            style_body_cell(cell, alt=alt, align='center')
        else:
            style_body_cell(cell, alt=alt, align='left')

    # AMC Seat Map URL column (last column, no csv_col)
    url_col_idx = len(SHOW_HEADERS)
    url_cell = ws3.cell(row=excel_row, column=url_col_idx, value=None)
    style_body_cell(url_cell, alt=alt, align='left')

# ── Sheet 4: Raw ──────────────────────────────────────────────────────────────
ws4 = wb.create_sheet('Raw')
write_headers(ws4, raw_cols)

raw_wide = {
    'theatre_name': 32, 'movie_title': 28, 'notes': 50,
    'polymarket_market': 36, 'auditorium_type': 28, 'check_time': 26
}
for i, col in enumerate(raw_cols, 1):
    set_col_width(ws4, i, raw_wide.get(col, 18))

for r_idx, (_, row) in enumerate(df_raw.iterrows()):
    excel_row = r_idx + 2
    alt = r_idx % 2 == 1
    for c_idx, col in enumerate(raw_cols, 1):
        val = row[col]
        if isinstance(val, float) and pd.isna(val):
            val = None
        cell = ws4.cell(row=excel_row, column=c_idx, value=val)
        style_body_cell(cell, alt=alt, align='left')

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Sheet 1 (Summary by Movie): {len(df_summary)} rows")
print(f"Sheet 2 (By Theatre): {len(df_theatre)} rows")
print(f"Sheet 3 (By Showtime): {len(df_byshow)} rows")
print(f"Sheet 4 (Raw): {len(df_raw)} rows")
